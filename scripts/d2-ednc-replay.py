#!/usr/bin/env python3
"""Meter Verse Billing Engine — EDNC Commercial Electricity Replay (Phase C/D2)"""

import sys, os, csv
from collections import Counter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(BASE, 'reference', 'collection-system'))
import tools.backup_manager
tools.backup_manager.backup = lambda: None
tools.backup_manager.run_backup = lambda: None
from app import create_app
from app.models import Tariff

import openpyxl

EXCEL_DIR = os.path.join(BASE, 'reference', 'meter-department', 'EDNC Migration 01 TO 05 - 2026')
OUTPUT_CSV = os.path.join(BASE, 'reports', 'd2-replay-results.csv')

BANDS = [(0, 100), (100, 250), (250, 600), (600, 1000), (1000, 999999)]
CSRV_BY_BAND = [5, 15, 20, 25, 40]
STAMP_RATE = 0.032
MIN_CHARGE = 9.10

# V1 rates (Jan-Mar): 0.85 / 1.68 / 2.20 / 2.27 / 2.33
# V1 M01 anomaly: no 2.27 tier — use 2.20 for 601-1000 band
RATES_V1_NORMAL = [0.85, 1.68, 2.20, 2.27, 2.33]
RATES_V1_M01 =    [0.85, 1.68, 2.20, 2.20, 2.33]
RATES_V2 = [1.62, 2.16, 2.64, 2.74, 2.79]

def get_band_index(cons):
    for i, (lo, hi) in enumerate(BANDS):
        if lo <= cons < hi:
            return i
    return len(BANDS) - 1

def calc_ednc_bill(cons, month):
    band_idx = get_band_index(cons)
    if month == 1:
        rates = RATES_V1_M01  # M01 anomaly
    elif month <= 3:
        rates = RATES_V1_NORMAL
    else:
        rates = RATES_V2

    rate = rates[band_idx]
    cc = round(cons * rate, 2)
    stamp = round(cons * STAMP_RATE, 2)
    csrv = CSRV_BY_BAND[band_idx]
    subtotal = round(cc + stamp + csrv, 2)
    total = max(subtotal, MIN_CHARGE)
    return rate, cc, stamp, csrv, total, band_idx

def main():
    app = create_app()
    app.app_context().push()

    # Verify DB tariffs exist
    v1 = Tariff.query.filter_by(name='EDNC Commercial V1').count()
    v2 = Tariff.query.filter_by(name='EDNC Commercial V2').count()
    print(f'DB tariffs: V1={v1} tiers, V2={v2} tiers')

    results = []
    total_rows = 0
    electricity_rows = 0
    water_rows = 0
    other_rows = 0

    for month_num in range(1, 6):
        fn = os.path.join(EXCEL_DIR, f'E & W EDNC {month_num:02d}-2026.xlsx')
        wb = openpyxl.load_workbook(fn, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
        total_rows += len(rows)
        print(f'Month {month_num:02d}: {len(rows)} rows')

        for r in rows:
            bill_no = r[0]
            meter_serial = str(r[2] or '').strip()
            meter_type = str(r[3] or '').strip().upper()
            unit_number = str(r[4] or '').strip()
            cons_raw = r[5]
            tax_raw = r[6]
            fees_raw = r[7]
            csrv_raw = r[8]
            admin_raw = r[9]
            total_raw = r[10]

            if cons_raw is None or (isinstance(cons_raw, str) and cons_raw.strip() == ''):
                other_rows += 1
                continue

            if meter_type != 'ELECTRICITY':
                water_rows += 1 if meter_type == 'WATER' else 0
                continue

            cons = float(cons_raw)
            tax_actual = float(tax_raw or 0)
            csrv_actual = float(csrv_raw or 0)
            fees_actual = float(fees_raw or 0)
            admin_actual = float(admin_raw or 0)
            total_actual = float(total_raw or 0)

            electricity_rows += 1
            if total_actual <= 0:
                # Zero-total rows: kFactor/estimate errors — record with 0 bill
                rate, cc_calc, stamp_calc, csrv_calc, total_calc, band_idx = 0, 0, 0, 0, 0, get_band_index(cons)
                month_str = f'{month_num:02d}'
                results.append({
                    'month': month_str, 'meter': meter_serial, 'unit': unit_number, 'type': meter_type,
                    'cons': round(cons, 3), 'rate': rate, 'band': band_idx,
                    'actual_consumption_charge': 0, 'actual_tax': 0, 'actual_csrv': 0, 'actual_total': 0,
                    'expected_cc': 0, 'expected_stamp': 0, 'expected_csrv': 0, 'expected_total': 0,
                    'total_diff': 0, 'pct_error': 0, 'status': 'ZERO',
                })
                continue

            rate, cc_calc, stamp_calc, csrv_calc, total_calc, band_idx = calc_ednc_bill(cons, month_num)
            month_str = f'{month_num:02d}'

            total_diff = round(total_actual - total_calc, 2)
            cc_diff = round((total_actual - csrv_actual - tax_actual - fees_actual - admin_actual) - cc_calc, 2)
            pct = abs(total_diff / total_calc * 100) if total_calc else 0

            if abs(total_diff) < 0.01:
                status = 'MATCH'
            elif abs(total_diff) <= 1.0:
                status = 'CLOSE'
            elif abs(total_diff) <= 5.0 and pct <= 5:
                status = 'OK_MIN'
            elif abs(total_diff) <= 20.0:
                status = 'MARGINAL'
            else:
                status = 'MISMATCH'

            results.append({
                'month': month_str,
                'meter': meter_serial,
                'unit': unit_number,
                'type': meter_type,
                'cons': round(cons, 3),
                'rate': rate,
                'band': band_idx,
                'actual_consumption_charge': round(total_actual - csrv_actual - tax_actual - fees_actual - admin_actual, 2),
                'actual_tax': tax_actual,
                'actual_csrv': csrv_actual,
                'actual_total': total_actual,
                'expected_cc': cc_calc,
                'expected_stamp': stamp_calc,
                'expected_csrv': csrv_calc,
                'expected_total': total_calc,
                'total_diff': total_diff,
                'pct_error': round(pct, 2),
                'status': status,
            })

    fieldnames = ['month', 'meter', 'unit', 'type', 'cons', 'rate', 'band',
                  'actual_consumption_charge', 'actual_tax', 'actual_csrv', 'actual_total',
                  'expected_cc', 'expected_stamp', 'expected_csrv', 'expected_total',
                  'total_diff', 'pct_error', 'status']
    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(results)

    statuses = Counter(r['status'] for r in results)
    diffs = [r['total_diff'] for r in results]
    avg_diff = sum(diffs) / len(diffs) if diffs else 0
    max_abs = max(abs(d) for d in diffs) if diffs else 0
    zero = sum(1 for d in diffs if abs(d) < 0.01)

    print(f'\n=== REPLAY SUMMARY ===')
    print(f'Total Excel rows: {total_rows}')
    print(f'Electricity rows: {electricity_rows}')
    print(f'Water rows:       {water_rows}')
    print(f'Other/skipped:    {other_rows}')
    print(f'Output rows:      {len(results)}')
    print(f'\nStatus counts: {dict(sorted(statuses.items()))}')
    print(f'Avg diff: {avg_diff:.4f}')
    print(f'Max abs diff: {max_abs:.2f}')
    print(f'Zero-diff (MATCH): {zero}/{len(results)}')
    print(f'\nOutput: {OUTPUT_CSV}')

if __name__ == '__main__':
    main()
