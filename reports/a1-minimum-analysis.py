"""Phase A-1: Solar Minimum Charge Root Cause Analysis"""
import sys, os
from pathlib import Path
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
import sqlite3
sys.stdout.reconfigure(encoding='utf-8')  # type: ignore

try:
    from openpyxl import load_workbook
except ImportError:
    os.system("pip install openpyxl -q")
    from openpyxl import load_workbook

BASE = Path(r'D:\meter\Meter\reference\collection-system')
OUT = Path(r'D:\meter\Meter\reports')

# ================================================================
# LOAD ALL DATA
# ================================================================
print("=== PHASE A-1: SOLAR MINIMUM CHARGE ANALYSIS ===")

# -- Solar invoices --
wb = load_workbook(str(BASE / 'Solar_Invoices_Import.xlsx'), read_only=True, data_only=True)
ws = wb['Solar Invoices']
invoices = []
next(ws.iter_rows(min_row=1, max_row=2))
for row in ws.iter_rows(min_row=3, values_only=True):
    if row[0] is None: continue
    invoices.append({
        'serial': str(row[0]).strip(),
        'month': str(row[1]),
        'amount': float(row[2] or 0),
        'inv_no': str(row[4] or ''),
        'tag': str(row[5] or ''),
        'info': str(row[6] or ''),
        'notes': str(row[7] or ''),
    })
wb.close()
print(f"Loaded {len(invoices)} invoices")

# -- Solar customers --
wb2 = load_workbook(str(BASE / 'Solar_Customers_For_Import.xlsx'), read_only=True, data_only=True)
ws2 = wb2['Customers']
customers = {}
for row in ws2.iter_rows(min_row=2, values_only=True):
    if row[0] is None: continue
    serial = str(row[8] or '').strip()
    if serial:
        customers[serial] = {
            'name': row[3], 'project': row[2], 'unit': str(row[17] or ''),
            'unit_type': row[18], 'customer_info': str(row[23] or ''),
        }
wb2.close()
print(f"Loaded {len(customers)} customers")

# -- Tariffs from DB --
conn = sqlite3.connect(str(BASE / 'instance' / 'collection.db'))
c = conn.cursor()
c.execute("SELECT id, name, type, mode, from_value, to_value, rate, flat_rate, version, valid_from, valid_to, status FROM tariff ORDER BY type, from_value")
tariffs = []
for row in c.fetchall():
    tariffs.append({
        'id': row[0], 'name': str(row[1] or ''), 'type': row[2],         'mode': row[3],
        'from': row[4], 'to': row[5], 'rate': row[6], 'flat_rate': row[7],
        'version': row[8], 'valid_from': row[9],
        'valid_to': row[10], 'status': str(row[11] or ''),
    })
print(f"Loaded {len(tariffs)} tariffs from collection.db")

# Additional tariff tables
c.execute("SELECT name FROM sqlite_master WHERE type='table'")
tables = [t[0] for t in c.fetchall()]
for tname in ['tariff_charge', 'tariff_charge_detail', 'tariff_version']:
    if tname in tables:
        c.execute(f"SELECT * FROM \"{tname}\"")
        cols = [d[0] for d in c.description]
        rows = c.fetchall()
        print(f"{tname}: {len(rows)} rows, columns: {cols}")
        for r in rows[:10]:
            print(f"  {r}")
conn.close()

# ================================================================
# PHASE 1: Identify ALL invoices with minimum charges
# ================================================================
print("\n=== PHASE 1: Affected Invoices ===")

# Group invoices by month, find distinct minimum amounts
by_month = {}
for inv in invoices:
    ym = inv['month']
    if ym not in by_month:
        by_month[ym] = []
    by_month[ym].append(inv['amount'])

# Find the "minimum" per customer group
by_serial_month = {}
for inv in invoices:
    key = (inv['serial'], inv['month'])
    by_serial_month[key] = inv['amount']

# For each customer, find their minimum
cust_mins = {}
for inv in invoices:
    s = inv['serial']
    if s not in cust_mins or inv['amount'] < cust_mins[s]:
        cust_mins[s] = inv['amount']

# Find all invoices that are at or near their customer's minimum
# Also find ALL unique invoice amounts to see the full distribution
from collections import Counter
amt_dist = Counter(inv['amount'] for inv in invoices)
print("\n--- Invoice Amount Distribution (Top 30) ---")
for amt, cnt in amt_dist.most_common(30):
    print(f"  {amt:>8.2f} EGP: {cnt:>4} invoices")

# Find all unique amounts under 50 EGP (minimum range)
print("\n--- All Invoice Amounts < 50 EGP ---")
small_amts = sorted([a for a in amt_dist if a < 50])
for amt in small_amts:
    # Find which months had this amount
    months_with = sorted(set(inv['month'] for inv in invoices if inv['amount'] == amt))
    serials_with = set(inv['serial'] for inv in invoices if inv['amount'] == amt)
    print(f"  {amt:>6.2f} EGP: {amt_dist[amt]:>4} invoices, months: {months_with[0]} → {months_with[-1]}, serials: {len(serials_with)}")

# ================================================================
# DEEP DIVE: Track 36.10 vs 9.10 by year
# ================================================================
print("\n\n=== MINIMUM CHARGE BY YEAR ===")
for year in ['2021', '2022', '2023', '2024', '2025', '2026']:
    year_invs = [inv for inv in invoices if inv['month'].startswith(year)]
    if not year_invs:
        continue
    amounts = [inv['amount'] for inv in year_invs]
    min_amt = min(amounts)
    max_amt = max(amounts)
    # Count how many are at minimum
    at_min = sum(1 for a in amounts if a <= min_amt * 1.05)
    print(f"  {year}: {len(year_invs)} invoices, min={min_amt:.2f}, max={max_amt:.2f}, at_min={at_min}")

# Find when the switch from 36.10 to 9.10 happened
print("\n\n=== TRANSITION TRACE ===")
transition_invs = [inv for inv in invoices if inv['amount'] in (9.10, 9.0, 36.10, 36.1)]
transition_invs.sort(key=lambda x: (x['serial'], x['month']))

# Group by serial, show the transition
by_serial_trans = {}
for inv in transition_invs:
    s = inv['serial']
    if s not in by_serial_trans:
        by_serial_trans[s] = []
    by_serial_trans[s].append(inv)

print("Customers with both 36.10 and 9.10 invoices (transition):")
for serial, invs in sorted(by_serial_trans.items()):
    amts = set(i['amount'] for i in invs)
    if len(amts) > 1:
        months_36 = [i['month'] for i in invs if i['amount'] == 36.10]
        months_9 = [i['month'] for i in invs if i['amount'] == 9.10]
        print(f"  {serial} ({customers.get(serial, {}).get('name', '?')[:20]}): 36.10={months_36}, 9.10={months_9}")

# ================================================================
# PHASE 2: Tariff Reconstruction
# ================================================================
print("\n\n=== PHASE 2: Tariff Reconstruction ===")
print(f"\nTariff table: {len(tariffs)} records")
print("\nAll tariffs by type:")
for ttype in set(t['type'] for t in tariffs):
    type_tariffs = [t for t in tariffs if t['type'] == ttype]
    print(f"\n  Type: {ttype} ({len(type_tariffs)} records)")
    for t in type_tariffs:
        print(f"    id={t['id']} name='{t['name']}' mode={t['mode']} "
              f"from={t['from']} to={t['to']} rate={t['rate']} "
              f"flat_rate={t['flat_rate']} "
              f"v={t['version']} from={t['valid_from']} to={t['valid_to']} status={t['status']}")

# ================================================================
# PHASE 3: Formula Reconstruction (try to calculate 36.10)
# ================================================================
print("\n\n=== PHASE 3: Formula Reconstruction ===")

# Try different fee combinations to arrive at 36.10
# Known: 36.10 = X + ESU(3). So X = 33.10
# For zero consumption:
# Total = CustomerService + TV(0) + Gov(0) + ConsFee(0) + ESU(3) + Jan(0 or 3)
# If Jan: 36.10 = CS + 0 + 0 + 0 + 3 + 3 = CS + 6 → CS = 30.10
# If not Jan: 36.10 = CS + 0 + 0 + 0 + 3 + 0 = CS + 3 → CS = 33.10

# Let's verify: 2021-01 (January) invoices
jan_2021_invs = [inv for inv in invoices if inv['month'] == '2021-01' and inv['amount'] == 36.10]
feb_2021_invs = [inv for inv in invoices if inv['month'] == '2021-02' and inv['amount'] == 36.10]

print(f"Jan 2021 @ 36.10: {len(jan_2021_invs)} invoices")
print(f"Feb 2021 @ 36.10: {len(feb_2021_invs)} invoices")
print(f"Both same amount → Jan fee (3 EGP) is INCLUDED in the minimum")
print(f"So: 36.10 = CS + TV(0) + Gov(0) + ConsFee(0) + ESU(3) + Jan(3)")
print(f"  → CS = 36.10 - 6 = 30.10")

# Now check if the non-Jan minimum also shows 36.10
# If Jan fee is included, then non-Jan should be different
# Feb 2021 IS NOT Jan, so if 36.10 = CS + 6 and Feb = CS + 3 = 33.10
# But Feb also = 36.10 → The Jan fee (3 EGP) is NOT always applied
# OR: the minimum is different

# Re-examine: let me look at which customers have 36.10 in Jan vs other months
print("\nChecking: are ALL months showing 36.10, or just some?")
months_with_36 = sorted(set(inv['month'] for inv in invoices if inv['amount'] == 36.10))
print(f"Months with 36.10 EGP minimum: {months_with_36}")

# Check: does 36.10 appear in 2022?
print("\n2022 months:")
for m in months_with_36:
    if m.startswith('2022'):
        print(f"  {m}")

# 9.10 analysis
print("\n\n--- 9.10 EGP Minimum Analysis ---")
feb_2022_9 = [inv for inv in invoices if inv['month'] == '2022-02' and abs(inv['amount'] - 9.10) < 0.01]
print(f"Feb 2022 @ 9.10: {len(feb_2022_9)} invoices")
# 9.10 = 6.10 + 3(ESU) or 9.10 = 9.0 + 0.10(Gov) or...
print("9.10 = 9.0 + 0.10 (governmental fee for 0 consumption)")
print("9.10 = 6.10 + 3.00 (ESU)")
print("9.10 = 0.48*0 (consumption) + 9.0 (CS) + 0.0 (TV) + 0.0 (Gov) + 0.0 (CF) + 0.0 (ESU?) + 0 (Jan)")

# Let's look at the Commercial sheet more carefully - it has different CS fees
# Commercial: CS = 20 EGP (not 9)
# The 9.0 could be a minimum CS fee for a particular category
# The 30.10 or 33.10 could be a different CS bracket

# Check the invoice_calculation_2020.xlsx for ALL categories' Customer Service fees
print("\n\n--- Customer Service Fees from invoice_calculation_2020.xlsx ---")
wb3 = load_workbook(str(BASE / 'New folder' / 'invoice_calculation_2020.xlsx'), read_only=True, data_only=True)
for sn in wb3.sheetnames:
    ws3 = wb3[sn]
    print(f"\n  Sheet: '{sn}' ({ws3.max_row} rows)")
    for row in ws3.iter_rows(min_row=1, max_row=min(25, ws3.max_row or 0), values_only=True):
        vals = [v for v in row]
        print(f"    {vals}")
wb3.close()

# ================================================================
# PHASE 4: Legacy Rule Discovery
# ================================================================
print("\n\n=== PHASE 4: Legacy Rule Discovery ===")

# Check the sbill Palm Hills billing schema for tariff rules
try:
    sbill_schema = Path(r'D:\meter\Meter\reference\sbill\OctoberBilling-Complete\01_database\PalmHills_Billing_FullSchema.sql')
    if sbill_schema.exists():
        with open(sbill_schema, 'r', encoding='utf-8') as f:
            content = f.read()
        # Find tariff-related tables
        for keyword in ['tariff', 'fee', 'charge', 'minimum', 'service', 'customer_service']:
            import re
            idx = content.lower().find(keyword)
            if idx >= 0:
                # Print 200 chars around each match
                start = max(0, idx - 50)
                end = min(len(content), idx + 200)
                print(f"\n--- '{keyword}' found at position {idx} ---")
                print(content[start:end])
except Exception as e:
    print(f"  Error reading sbill schema: {e}")

# Check for any minimum charge rules in collection system code
print("\n\n--- Checking reference code for minimum charge logic ---")
charge_engine = Path(r'D:\meter\Meter\reference\collection-system\app\charge_engine.py')
if charge_engine.exists():
    with open(charge_engine, 'r', encoding='utf-8') as f:
        content = f.read()
    # Look for minimum/fixed/static charge patterns
    for kw in ['ZERO', 'minimum', 'STATIC', 'fixed', 'service', 'flat_amount', '36.10', 'min']:
        import re
        matches = re.finditer(f'.{{0,80}}{kw}.{{0,80}}', content, re.IGNORECASE)
        for m in matches:
            print(f"  [{kw}] {m.group().strip()}")

# Check routes for solar-specific logic
routes_admin = Path(r'D:\meter\Meter\reference\collection-system\app\routes_admin.py')
if routes_admin.exists():
    with open(routes_admin, 'r', encoding='utf-8') as f:
        content = f.read()
    for kw in ['solar', 'minimum', 'service_fee', '36.10', 'admin_fee']:
        import re
        matches = re.finditer(f'.{{0,80}}{kw}.{{0,80}}', content, re.IGNORECASE)
        for m in matches:
            print(f"  [admin/{kw}] {m.group().strip()}")

# ================================================================
# PHASE 5: Impact Assessment
# ================================================================
print("\n\n=== PHASE 5: Impact Assessment ===")

# Count invoices at each minimum level
amt_36_invs = [inv for inv in invoices if abs(inv['amount'] - 36.10) < 0.01]
amt_9_invs = [inv for inv in invoices if abs(inv['amount'] - 9.10) < 0.01]
amt_9plain = [inv for inv in invoices if abs(inv['amount'] - 9.00) < 0.01]

print(f"Invoices @ 36.10 EGP: {len(amt_36_invs)}")
print(f"  Unique customers: {len(set(i['serial'] for i in amt_36_invs))}")
print(f"  Total value: {sum(i['amount'] for i in amt_36_invs):.2f} EGP")
print(f"  If corrected to 9.10: {sum(9.10 for i in amt_36_invs):.2f} EGP")
print(f"  Difference: {sum(i['amount'] for i in amt_36_invs) - sum(9.10 for i in amt_36_invs):.2f} EGP")

print(f"\nInvoices @ 9.10 EGP: {len(amt_9_invs)}")
print(f"  Unique customers: {len(set(i['serial'] for i in amt_9_invs))}")
print(f"  Total value: {sum(i['amount'] for i in amt_9_invs):.2f} EGP")

print(f"\nInvoices @ 9.00 EGP: {len(amt_9plain)}")
print(f"  Unique customers: {len(set(i['serial'] for i in amt_9plain))}")

# Find ALL distinct invoice amounts that appear as "minimum" for each customer
print("\n\n--- Per-Customer Minimum Amounts ---")
for serial in sorted(cust_mins.keys())[:15]:
    cust_invs = [inv for inv in invoices if inv['serial'] == serial]
    amounts = sorted(set(inv['amount'] for inv in cust_invs))
    min_amt = min(amounts)
    # Find how many months are at this minimum
    at_min = sum(1 for inv in cust_invs if inv['amount'] == min_amt)
    name = customers.get(serial, {}).get('name', '?')
    print(f"  {serial} ({str(name)[:20]}): min={min_amt}, all_amounts={amounts[:10]}{'...' if len(amounts)>10 else ''}, at_min_months={at_min}")

# ================================================================
# RECONSTRUCT: What fee structure gives 36.10 and 9.10?
# ================================================================
print("\n\n=== FEE STRUCTURE RECONSTRUCTION ===")
print("\nGoal: Find fee combination that produces 36.10 and 9.10")

# The 2021 residential sheet shows:
# CS: 9 | TV: 0 | Gov: 0.01 | Cons: 0 | ESU: 3 | Jan: 3 | Total: 15.01 (or ~15.09)
# But we see 36.10 and 9.10 in the data

# Hypothesis: 36.10 = commercial tariff with higher CS
# Commercial sheet shows CS = 20 EGP (not 9)
# 20 + TV(0) + Gov(0.01) + ConsFee(0) + ESU(3) + Jan(3) = 26.01
# Still not 36.10

# Hypothesis: 36.10 = different fee schedule entirely
# Let me look at: what if CS = 30.10 (for a higher consumption bracket)?
# The difference: 36.10 - 3(ESU) - 3(Jan) = 30.10 → CS for some bracket
# Non-Jan: 36.10 - 3(ESU) = 33.10 → CS = 33.10 for non-Jan... but that doesn't work
# Since Jan and Feb are BOTH 36.10, the Jan fee ISN'T always applied differently

# Wait — let me re-examine. Let me check if 36.10 appears in FEB 2021, MAR 2021, etc.
# If 36.10 appears in ALL months of 2021, then the minimum is truly 36.10 flat
print("\n2021 monthly breakdown:")
for m in ['2021-01', '2021-02', '2021-03', '2021-04', '2021-05', '2021-06',
          '2021-07', '2021-08', '2021-09', '2021-10', '2021-11', '2021-12']:
    minvs = [inv for inv in invoices if inv['month'] == m]
    if minvs:
        amounts = [inv['amount'] for inv in minvs]
        min_a = min(amounts)
        at_min = sum(1 for a in amounts if abs(a - min_a) < 0.01)
        print(f"  {m}: {len(minvs)} invoices, min={min_a:.2f}, at_min={at_min}")

print("\n2022 monthly breakdown:")
for m in ['2022-01', '2022-02', '2022-03', '2022-04', '2022-05', '2022-06',
          '2022-07', '2022-08', '2022-09', '2022-10', '2022-11', '2022-12']:
    minvs = [inv for inv in invoices if inv['month'] == m]
    if minvs:
        amounts = [inv['amount'] for inv in minvs]
        min_a = min(amounts)
        at_min = sum(1 for a in amounts if abs(a - min_a) < 0.01)
        print(f"  {m}: {len(minvs)} invoices, min={min_a:.2f}, at_min={at_min}")
